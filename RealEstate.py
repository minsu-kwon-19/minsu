import math
from posixpath import split
import requests
from bs4 import BeautifulSoup
import json
import re
import math
import sys
import time, datetime
import csv

import numpy as np
import pandas as pd
import seaborn as sns
import timeit
import collections
from dataclasses import dataclass
from typing import NamedTuple

import openpyxl as op 

#
#   startStr부터 ~ endStr 사이의 문자열 반환
#
def getStrBetweenAnB(fullStr, startStr, endStr):
    startStrLen = len(startStr)    

    # find start index
    startIdx = fullStr.find(startStr) + startStrLen

    # find end index
    endIdx = fullStr.find(endStr, startIdx)

    return fullStr[startIdx:endIdx]

#
#    Get Response
#
def getRes(url, headers):
    res = requests.get(url, headers = headers)
    time.sleep(1)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")
        
    return res

#
#    해당 층수 리턴.  
#
def getFloorInfo(floorInfo):
    result = floorInfo.split('/')
    return str(result[0])

#
#    string으로 저장된 가격을 int 값으로 변경하여 반환.
#
def getPriceInfo(price):
    list = price.split('억 ')
    result = 0
    
    if len(list) == 2: # ex) 3억 6,000
        result = int(list[0]) * 10000 + int(list[1].replace(',', ''))
    else: # ex) 3억
        result = int(list[0].replace('억','')) * 10000
        
    return int(result)

#
#    공급면적을 기준으로 각 면적당 최소값만 리턴.
#
def getMinVal(aptList):
    
    aptList.sort()
    rmEleList = [] # 삭제할 element를 임시 저장하는 리스트.
                
    criteria = 0.0 # 공급면적 기준 값을 저장하는 변수. float.
                
    for val in range(0, len(aptList)):
        if val == 0:    # 단지에서 공급면적 기준으로 가장 저렴한 가격.
            criteria = float(aptList[val].spc1)
            continue

        if float(aptList[val].spc1) == float(criteria):
            rmEleList.append(aptList[val])
        else:
            criteria = float(aptList[val].spc1)
                            
    # 최저값 제외하고 지우는 과정.
    for rm in rmEleList:
        aptList.remove(rm)
        
    return aptList

#
#    저층과 5층 이하 매물을 없애주고 반환하는 기능
#
def removeLowFloor(apt, aptList):
    
    # 층 정보, 가격 정보
    floorInfo = getFloorInfo(apt['flrInfo'])
    priceInfo = getPriceInfo(apt['prcInfo'])
    
    if floorInfo == "저":
        return
    elif  floorInfo == '중' or floorInfo == '고':
        aptList.append(RealEstateInfo(f"{apt['atclNm']}", float(apt['spc1']), float(apt['spc2']), priceInfo, f"{floorInfo}"))                                        
    elif int(floorInfo) < 5:
        return
    else:
        aptList.append(RealEstateInfo(f"{apt['atclNm']}", float(apt['spc1']), float(apt['spc2']), priceInfo, f"{floorInfo}"))         

#
#    real estate struct
#
class RealEstateInfo(NamedTuple):
    name:str
    spc1:float
    spc2:float
    priceInfo:int
    floorInfo:str

        
#
#       Main
#
headers = {'User-Agent': "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 1.1.4322; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729; Browzar)"}
keyword1 = "송파구"
excelFilePath = "./시세표 송파구.xlsm"

# 구에서 동 cortano 얻어오기.
guUrl = "https://m.land.naver.com/search/result/" + keyword1    # 구단위
resGuArea = getRes(guUrl, headers)
strGuResult = getStrBetweenAnB(resGuArea.text, "filter: {", "},")
guCortarNo = getStrBetweenAnB(strGuResult, "cortarNo: '","',")

# 해당 구의 행정동 리스트 뽑아오기.
dongUrl = f"https://m.land.naver.com/map/getRegionList?cortarNo={guCortarNo}&mycortarNo={guCortarNo}"
resDong = getRes(dongUrl, headers)
dongJsonObject = json.loads(resDong.text)
dongArray = dongJsonObject["result"]["list"]

for dong in dongArray:
    keyword2 = dong['CortarNm']    # 송파동, 잠실동 등.. 동 이름 keyword

    url = "https://m.land.naver.com/search/result/" + keyword1 + keyword2

    resArea = getRes(url, headers)

    strResult = getStrBetweenAnB(resArea.text, "filter: {", "},")

    lat = getStrBetweenAnB(strResult, "lat: '","',")
    lon = getStrBetweenAnB(strResult, "lon: '","',")
    z = getStrBetweenAnB(strResult, "z: '","',")
    cortarNo = getStrBetweenAnB(strResult, "cortarNo: '","',")
    searchType = "APT"

    rletTpCds = getStrBetweenAnB(strResult, "rletTpCds: '","',")
    tradTpCds = getStrBetweenAnB(strResult, "tradTpCds: '","',")

    # 지도에서 단지들 url
    onMapURL = f"https://m.land.naver.com/cluster/clusterList?cortarNo={cortarNo}&rletTpCd={searchType}&tradTpCd=A1%3AB1%3AB2&z={z}&lat={lat}&lon={lon}&btm=37.4756089&lft=127.099033&addon=COMPLEX&bAddon=COMPLEX&isOnlyIsale=false"
    resMap = getRes(onMapURL, headers)
    mapJsonObject = json.loads(resMap.text)
    mapArray = mapJsonObject["data"]["COMPLEX"]

    wb = op.load_workbook(excelFilePath, data_only = 'True', read_only=False, keep_vba=True) #worksheet 객체 생성

    ws = wb[f"{keyword2}"]

    print(f"workSheet name : {ws}")

    # 매물 정보
    for val in mapArray:
        # 매물 url
        # 20개 이상일땐 page 0부터시작
        # 20개 이하일땐 page 1부터 시작
        ipage = int(int(val['count'])/20) # 20개씩으로 끊어짐.
        if int(int(val['count'])/20) == 0:
            ipage = ipage + 2    
        else:
            ipage = ipage + 1

        # 각 지도에서 20개의 매물이 합쳐져 보이는거 체크.
        for pageCnt in range(1, ipage):
            # 매물 20개가 넘는 위치들 카운트 하기 위함.
            offeringsURL = f"https://m.land.naver.com/cluster/ajax/complexList?itemId={val['lgeo']}&lgeo={val['lgeo']}&rletTpCd=APT&tradTpCd=A1:B1:B2&z={z}&lat={val['lat']}&lon={val['lon']}&btm=37.4404697&lft=127.1156627&top=37.5303033&rgt=127.1285373&cortarNo={cortarNo}&isOnlyIsale=false&sort=readRank&page={pageCnt}"
            resOfferings = getRes(offeringsURL, headers)

            offeringsJsonObject = json.loads(resOfferings.text)
            offeringsArray = offeringsJsonObject["result"]

            # 지역 전체 매물 array
            for v in offeringsArray:
                if int(v['totHsehCnt']) > 300:    # total Households
                    print(f"{v['hscpNm']}, {v['dealCnt']}, {v['leaseCnt']}")

                    # 아파트 매물 url
                    num = int(v['dealCnt']) + int(v['leaseCnt'])                
                    aptPage =  int(num / 20) + 2

                    myDealList = []
                    myLeaseList = []

                    # 매매, 전세 페이지 처리. # 20개 이상일 때 page 증가됨.
                    for aptCnt in range(1, aptPage):
                        aptUrl = f'https://m.land.naver.com/complex/getComplexArticleList?hscpNo={v["hscpNo"]}&cortarNo=1171010800&tradTpCd=A1:B1&order=point_&showR0=N&page={aptCnt}'
                        resApt = getRes(aptUrl, headers)

                        aptJsonObject = json.loads(resApt.text)
                        aptArray = aptJsonObject['result']['list']

                        # APT 단지 매물 검색
                        # 층수 "flrInfo":"11/15" "저/15", "고/15"
                        # 공급/전용 "spc1":"103.12","spc2":"84.7"
                        # 매매와 전세 리스트 따로 저장하기. 
                        for apt in aptArray:  
                            if apt['tradTpNm'] == "매매":
                                removeLowFloor(apt, myDealList)
                            elif apt['tradTpNm'] == "전세":
                                removeLowFloor(apt, myLeaseList)

                    # 매매와 전세가격중 최소값만 저장.
                    # 엑셀 데이터에 입력해야함.
                    myDealList = getMinVal(myDealList)
                    myLeaseList = getMinVal(myLeaseList)

                    # input data.
                    for cnt in range(6, ws.max_row + 1):    # sheet range
                        for dealApt in myDealList:  # APT range
                            if ws[f'A{cnt}'].value == dealApt.name and ws[f'D{cnt}'].value == round(dealApt.spc1, 1) and ws[f'F{cnt}'].value == round(dealApt.spc2,1):
                                ws[f'I{cnt}'] = str(dealApt.priceInfo)
                                ws[f'J{cnt}'] = str(dealApt.floorInfo)

                        for leaseApt in myLeaseList:
                            if ws[f'A{cnt}'].value == leaseApt.name and ws[f'D{cnt}'].value == round(leaseApt.spc1, 1) and ws[f'F{cnt}'].value == round(leaseApt.spc2, 1):
                                ws[f'K{cnt}'] = str(leaseApt.priceInfo)
                                ws[f'L{cnt}'] = str(leaseApt.floorInfo)

                    # save APT info
                    wb.save(f"{excelFilePath}")


                        